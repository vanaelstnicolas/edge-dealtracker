from datetime import date
from io import BytesIO
from typing import Literal
from typing import Any
import unicodedata

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook

from app.api.deps.auth import get_current_user
from app.config import settings
from app.repositories.in_memory import store
from app.schemas.deal import DealCreate, DealRead, DealStatus, DealUpdate
from app.services.rate_limit import enforce_rate_limit

router = APIRouter()


def _is_admin(current_user: dict[str, Any]) -> bool:
    app_metadata = current_user.get("app_metadata")
    if not isinstance(app_metadata, dict):
        return False

    role = app_metadata.get("role")
    if isinstance(role, str) and role.lower() == "admin":
        return True

    roles = app_metadata.get("roles")
    if isinstance(roles, list):
        return any(isinstance(item, str) and item.lower() == "admin" for item in roles)

    return False


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value.strip().lower())
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    compact = " ".join(ascii_only.split())
    return compact


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _infer_status(*chunks: str) -> DealStatus:
    text = " ".join(chunk.lower() for chunk in chunks if chunk)
    lost_markers = ("cloture", "cloturee", "perdu", "perdue", "reponse negative", "abandon")
    won_markers = ("gagne", "gagnee", "signature", "signe", "signee")

    if any(marker in text for marker in lost_markers):
        return DealStatus.lost
    if any(marker in text for marker in won_markers):
        return DealStatus.won
    return DealStatus.active


def _extract_headers(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    header_aliases = {
        "company": {"cibles commerciales", "cible commerciale", "entreprise", "client"},
        "progress": {"avancement", "statut", "progression"},
        "action": {"actions commerciales", "actions", "action commerciale", "action"},
        "comments": {
            "autres actions / questions - commentaires",
            "autres actions/questions - commentaires",
            "commentaires",
            "notes",
        },
    }

    for row_index, row in enumerate(rows):
        normalized_cells = [_normalize_header(_to_text(cell)) for cell in row]
        positions: dict[str, int] = {}
        for field, aliases in header_aliases.items():
            for index, cell_value in enumerate(normalized_cells):
                if cell_value in aliases:
                    positions[field] = index
                    break
        if "company" in positions and "action" in positions:
            return row_index, positions

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Excel headers not recognized. Expected columns include 'Cibles commerciales' and 'Actions commerciales'.",
    )


@router.post("/import/excel")
def import_deals_excel(
    current_user: dict[str, Any] = Depends(get_current_user),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    owner_id = str(current_user["id"])
    enforce_rate_limit(
        bucket="deals_import",
        key=owner_id,
        limit=settings.deals_import_rate_limit_per_user,
        window_seconds=settings.rate_limit_window_seconds,
    )

    file_name = (file.filename or "").lower()
    if not file_name.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx files are supported")

    payload = file.file.read()
    if not payload:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")

    try:
        workbook = load_workbook(BytesIO(payload), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unable to parse Excel file") from exc

    worksheet: Any = workbook.active
    if worksheet is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel sheet is empty")
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel sheet is empty")

    header_row_index, positions = _extract_headers(rows)
    imported = 0
    skipped = 0
    errors: list[str] = []

    for row_number, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        company = _to_text(row[positions["company"]]) if positions.get("company") is not None and positions["company"] < len(row) else ""
        action = _to_text(row[positions["action"]]) if positions.get("action") is not None and positions["action"] < len(row) else ""
        progress = _to_text(row[positions["progress"]]) if positions.get("progress") is not None and positions["progress"] < len(row) else ""
        comments = _to_text(row[positions["comments"]]) if positions.get("comments") is not None and positions["comments"] < len(row) else ""

        if not company:
            skipped += 1
            continue

        description = progress or comments or action
        if len(description) < 2:
            skipped += 1
            errors.append(f"Row {row_number}: missing usable description")
            continue
        if len(action) < 2:
            action = "A qualifier"

        normalized_description = description[:2000]
        normalized_action = action[:240]
        normalized_company = company[:140]

        deal = DealCreate(
            company=normalized_company,
            description=normalized_description,
            action=normalized_action,
            deadline=date.today(),
            owner_id=owner_id,
            status=_infer_status(progress, action, comments),
        )
        store.create_deal(deal)
        imported += 1

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:20],
    }


@router.get("", response_model=list[DealRead])
def list_deals(
    current_user: dict[str, Any] = Depends(get_current_user),
    status_value: DealStatus | None = Query(default=None, alias="status"),
    owner_id: str | None = Query(default=None),
    scope: Literal["all", "active", "archived"] = Query(default="all"),
) -> list[DealRead]:
    user_id = str(current_user["id"])
    if owner_id is not None and owner_id != user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden owner scope")

    deals = store.list_deals(status=status_value, owner_id=user_id)
    if scope == "active":
        return [deal for deal in deals if deal.status == DealStatus.active]
    if scope == "archived":
        return [deal for deal in deals if deal.status in {DealStatus.won, DealStatus.lost}]
    return deals


@router.post("", response_model=DealRead, status_code=status.HTTP_201_CREATED)
def create_deal(payload: DealCreate, current_user: dict[str, Any] = Depends(get_current_user)) -> DealRead:
    user_id = str(current_user["id"])
    if payload.owner_id != user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden owner scope")
    return store.create_deal(payload)


@router.patch("/{deal_id}", response_model=DealRead)
def update_deal(deal_id: str, payload: DealUpdate, current_user: dict[str, Any] = Depends(get_current_user)) -> DealRead:
    user_id = str(current_user["id"])
    admin = _is_admin(current_user)

    visible_rows = store.list_deals(status=None, owner_id=None if admin else user_id)
    if all(deal.id != deal_id for deal in visible_rows):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Deal not found")

    if payload.owner_id is not None and payload.owner_id != user_id and not admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden owner scope")

    updated = store.update_deal(deal_id, payload)
    if updated is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Deal not found")
    return updated
